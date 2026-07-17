#!/usr/bin/env python3
"""
Parser des factures Excel du Lac Hôtel (dossier PRO FORMA 2026).

Lecture SEULE (openpyxl data_only : lit les valeurs en cache, ne recalcule
pas, ne sauvegarde jamais → les =TODAY() ne sont pas ré-évalués).

Le gabarit est stable mais les lignes de totaux FLOTTENT (leur ligne dépend
du nombre de prestations) : on les repère en scannant la colonne B, pas par
position fixe.

Sortie : factures.jsonl (une facture par ligne) + anomalies.json.
"""
import json
import os
import re
import unicodedata
from datetime import datetime

import openpyxl

RACINE = "/Volumes/MAX/_Lachotel (Site Refonte)/Facturation du Lac Hôtel/PRO FORMA 2026"
OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def norm(s):
    """Minuscule, sans accents, espaces normalisés (nbsp inclus)."""
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ").replace(" ", " ")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = re.sub(r"[^\d,.\-]", "", str(v)).replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def to_iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    return None


def type_document(c1):
    n = norm(c1)
    if n.startswith("avoir") or "note de credit" in n or "note d avoir" in n:
        return "AVOIR"
    if n.startswith("pro"):
        return "PROFORMA"
    if n.startswith("fact"):
        return "FACTURE"
    return None


def extraire_numero(c1):
    """« FACT 001 2026 » → 'FACT 001 2026' ; 'PRO-FORMA' → None."""
    n = norm(c1)
    if n.startswith("pro"):
        return None
    m = re.search(r"(fact|avoir)\D*(\d{1,5})\D*(\d{4})?", n)
    if m:
        return c1.strip()
    return None


def pax_entier(v):
    """« 12 + 1 TL » → 12 ; 42 → 42."""
    if isinstance(v, (int, float)):
        return int(v)
    m = re.match(r"\s*(\d+)", str(v or ""))
    return int(m.group(1)) if m else None


LABELS_HT = ("sous total ht", "sous-total ht", "total ht")
LABELS_TVA = ("tva 20", "tva20", "tva")
LABELS_VIGNETTE = ("vignette", "vingette")  # typo « vingette » incluse
LABELS_TTC = ("montant total ttc", "total ttc", "montant total")
LABELS_AVOIR = ("montant de l avoir", "montant de l'avoir")


def trouver_header(ws):
    """Ligne où A='Nombre' et B='Désignations'."""
    for r in range(15, 26):
        if norm(ws[f"A{r}"].value) == "nombre" and "designation" in norm(ws[f"B{r}"].value):
            return r
    return None


def parser_fichier(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    c1 = ws["C1"].value
    typ = type_document(c1)

    header = trouver_header(ws)
    # Détecter un fichier NON-facture (planning, fiche info) : pas de header et
    # pas de MONTANT TOTAL TTC.
    max_row = ws.max_row or 45

    # Repérer les lignes de totaux + la 1re ligne de total (fin des prestations)
    tot = {"ht": None, "tva": None, "vignette": None, "ttc": None, "avoir": None, "remise": None}
    ligne_premier_total = None
    for r in range(1, min(max_row, 60) + 1):
        b = norm(ws[f"B{r}"].value)
        e = ws[f"E{r}"].value
        if not b:
            continue
        matched = None
        if any(b == l or b.startswith(l) for l in LABELS_HT):
            tot["ht"] = to_num(e); matched = r
        elif any(l in b for l in LABELS_TTC):
            tot["ttc"] = to_num(e); matched = r
        elif any(l in b for l in LABELS_AVOIR):
            tot["avoir"] = to_num(e); matched = r
        elif b.startswith("remise"):
            tot["remise"] = to_num(e); matched = r
        elif any(b.startswith(l) for l in LABELS_TVA):
            tot["tva"] = to_num(e); matched = r
        elif any(l in b for l in LABELS_VIGNETTE):
            tot["vignette"] = to_num(e); matched = r
        if matched and matched > (header or 0) and ligne_premier_total is None:
            ligne_premier_total = matched

    est_facture = typ is not None and (tot["ttc"] is not None or tot["avoir"] is not None or header is not None)
    if not est_facture:
        wb.close()
        return None, "non-facture (planning / fiche info)"

    # Prestations : entre header+1 et le 1er total (ou vignette exclue si elle
    # est après). On garde toute ligne A/B/E numérique dans cette plage.
    lignes = []
    if header:
        fin = ligne_premier_total or (header + 20)
        for r in range(header + 1, fin):
            a = ws[f"A{r}"].value
            b = ws[f"B{r}"].value
            e = ws[f"E{r}"].value
            if b is None and e is None:
                continue
            nb = norm(b)
            # ne pas confondre une ligne total glissée dans la plage
            if any(x in nb for x in ("sous total", "sous-total", "tva ", "montant total", "montant de l")):
                continue
            lignes.append({
                "nombre": to_num(a),
                "designation": str(b).replace("\xa0", " ").strip() if b else "",
                "pu": to_num(ws[f"C{r}"].value),
                "nuitees": to_num(ws[f"D{r}"].value),
                "total": to_num(e),
                # vignette = hors TVA
                "soumis_tva": "vignette" not in nb and "vingette" not in nb,
            })

    # Bloc client
    def cell(ref):
        v = ws[ref].value
        return str(v).replace("\xa0", " ").strip() if isinstance(v, str) else v

    agence_folder = os.path.basename(os.path.dirname(path))
    fac = {
        "fichier": os.path.relpath(path, RACINE),
        "agence_dossier": agence_folder,
        "type": typ,
        "numero": extraire_numero(c1),
        "titre_brut": cell("C1"),
        "client": cell("C5") or cell("D12"),
        "agence_bloc": cell("C6"),
        "date_emission": to_iso(ws["D3"].value),  # =TODAY() : peu fiable
        "date_in": to_iso(ws["D13"].value),
        "date_out": to_iso(ws["D14"].value),
        "prestations_code": cell("D15"),
        "pax_brut": cell("D16"),
        "pax": pax_entier(ws["D16"].value),
        "nuits": to_num(ws["D17"].value),
        "lignes": lignes,
        "sous_total_ht": tot["ht"],
        "tva": tot["tva"],
        "vignette": tot["vignette"],
        "total_ttc": tot["ttc"],
        "remise": tot["remise"],
        "montant_avoir": tot["avoir"],
        "mention_paiement": cell("B37") or cell("B38") or cell("B39"),
    }
    wb.close()
    return fac, None


def main():
    fichiers = []
    for root, _, names in os.walk(RACINE):
        for n in names:
            if n.lower().endswith(".xlsx") and not n.startswith("._") and not n.startswith("~$"):
                fichiers.append(os.path.join(root, n))
    fichiers.sort()

    factures, anomalies = [], []
    for f in fichiers:
        try:
            fac, skip = parser_fichier(f)
            if skip:
                anomalies.append({"fichier": os.path.relpath(f, RACINE), "raison": skip})
                continue
            # contrôles de cohérence
            probs = []
            if fac["total_ttc"] is None and fac["montant_avoir"] is None:
                probs.append("aucun total")
            if fac["type"] != "AVOIR" and not fac["lignes"]:
                probs.append("aucune ligne")
            if fac["type"] != "AVOIR" and fac["total_ttc"]:
                somme = sum((l["total"] or 0) for l in fac["lignes"])
                # TTC = somme lignes + TVA (si lignes HT) — tolérance large, info seulement
                if somme and abs(somme) > 0:
                    fac["_somme_lignes"] = round(somme)
            if probs:
                fac["_anomalies"] = probs
                anomalies.append({"fichier": fac["fichier"], "raison": ", ".join(probs)})
            # Montant CA canonique : TTC si présent, sinon somme des lignes ;
            # avoir = négatif.
            if fac["type"] == "AVOIR":
                fac["montant_ca"] = -abs(fac["montant_avoir"]) if fac["montant_avoir"] else None
            else:
                fac["montant_ca"] = fac["total_ttc"] if fac["total_ttc"] is not None else (
                    round(sum((l["total"] or 0) for l in fac["lignes"])) or None)
            factures.append(fac)
        except Exception as e:
            anomalies.append({"fichier": os.path.relpath(f, RACINE), "raison": f"ERREUR: {e}"})

    with open(f"{OUT_DIR}/factures.jsonl", "w", encoding="utf-8") as fh:
        for fac in factures:
            fh.write(json.dumps(fac, ensure_ascii=False) + "\n")
    with open(f"{OUT_DIR}/anomalies.json", "w", encoding="utf-8") as fh:
        json.dump(anomalies, fh, ensure_ascii=False, indent=1)

    # Synthèse
    par_type = {}
    for f in factures:
        par_type[f["type"]] = par_type.get(f["type"], 0) + 1
    print(f"Fichiers .xlsx scannés : {len(fichiers)}")
    print(f"Documents extraits     : {len(factures)}  {par_type}")
    print(f"Anomalies / ignorés    : {len(anomalies)}")


if __name__ == "__main__":
    main()
